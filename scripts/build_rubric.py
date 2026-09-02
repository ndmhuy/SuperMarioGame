#!/usr/bin/env python3
"""Generate Group 52's self-assessment against the CS202 rubric.

One sheet, this group only. The source rubric workbook also carries a previous
cohort's marks for other groups; those are deliberately not reproduced here -
they are other people's grades and none of our business.

Each row states the mark we propose AND the evidence for it, so a grader can
check the claim instead of taking it. The 3D row proposes 0: no 3D renderer was
built, and the report says so in its known-gaps section rather than implying
otherwise.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROWS = [
    ("PlayerInputsMovementCollision", 20, 20,
     "Fixed 1/60s timestep with interpolated rendering; AABB collision through "
     "PhysicsEngine + CollisionResolver (0 real dynamic_cast after the R5 refactor); "
     "walk/run/jump/wall-jump/ground-pound/crouch-slide; coyote time and jump "
     "buffering at 6 frames each; full key rebinding for both players.",
     "features_list #1-#20; report 6.2, 8.2; verify_physics, verify_collision"),
    ("EnemyBehavior", 10, 10,
     "13 enemy types plus 3 colour variants, Boom Boom as mid-boss and Bowser with "
     "a multi-phase fight, bridge chop and lava death. 8 interchangeable movement "
     "strategies swapped at runtime - a stomped Paratroopa becomes a Koopa by "
     "exchanging its strategy object.",
     "features_list #21-#43; report 7.5; verify_enemies, verify_boss_fights"),
    ("PowerUpsItems", 10, 10,
     "13 item types; 5 base player forms (Small/Super/Fire/Cape/Mini) plus Star and "
     "Mega as stacking Decorators that forward through the wrapped state, so a Fire "
     "Flower survives a Star.",
     "features_list #24-#47; report 7.12; verify_powerups"),
    ("3 LevelCompletion", 15, 15,
     "3 themed levels (overworld, underground, castle) at 200 tiles wide, plus 3 "
     "sub-vaults, a bonus room, an SMB3-style world map, star coins, and a flagpole "
     "completion path. Endless Mode generates further levels with a solvability "
     "oracle.",
     "features_list #48-#67, #96; report 8.4; verify_level_data, verify_map_generator"),
    ("Sounds", 10, 10,
     "29 sound effects and 12 music tracks; surface-dependent footsteps; combo SFX "
     "pitch escalation; per-row menu cues; dynamic music layers responding to boss "
     "proximity and low timer. SoundManager subscribes to 20 event types.",
     "features_list #68-#77; report 7.11; verify_audio"),
    ("OOD", 10, 10,
     "Audited directly, not asserted: 0 public mutable data members in any class; "
     "ownership is unique_ptr throughout with 0 delete; every one of the 10 hierarchy "
     "roots has a virtual destructor. Trade-offs are documented as decisions with "
     "reasons rather than hidden - PlayingState's size, 12 singletons, friend classes.",
     "report 6.1-6.8 (SOLID walk-through, one principle per paragraph)"),
    ("DesignPatterns", 25, 25,
     "13 distinct GoF patterns across 16 report subsections, each written as problem "
     "-> naive alternative -> why this pattern -> what it cost. Includes patterns the "
     "spec never claimed (the editor's Command with undo/redo, Registry/Type Object, "
     "Composite, Adapter) and one deliberately REJECTED (Visitor / double dispatch, "
     "in favour of an ordered EntityCategory pair switch).",
     "report 7.1-7.16; verify_r21_entity_registry enforces the Factory's openness"),
    ("AI", 5, 5,
     "IAIPolicy::decide(AIObservation) -> AIAction as a policy seam, with "
     "HeuristicPolicy as the shipped baseline and AIController sensing/actuating. "
     "Reaches the player through a CPU opponent, Shadow Mario chase, and a 2P AI mode.",
     "features_list #78-#83; report 7.13, 14.1; verify_multiplayer_ai"),
    ("MultiplePlayers", 5, 5,
     "Local 2P versus and co-op, Shadow Chase against a recorded ghost, and a CPU "
     "opponent. Includes a survivor camera and eliminated-player badges when one "
     "player is out.",
     "features_list #84-#90, #117-#118; report 8.6; verify_multiplayer_ai"),
    ("3D Game", 5, 0,
     "NOT ATTEMPTED. The project is deliberately 2D; no 3D renderer was built. "
     "Proposing 0 rather than arguing for partial credit.",
     "report 13 (known gaps) states this explicitly"),
]

def build(out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Group52"

    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F5496")
    thin = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    wrap = Alignment(vertical="top", wrap_text=True)

    ws["A1"] = "CS202 Programming Systems - Final Project"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Group 52, APCS K25 - Super Mario Game | Nguyen Dinh Minh Huy (25125083), "
                "Tran Gia Huy (25125084)")
    ws["A3"] = "Self-assessment against the rubric. Every proposed mark carries its evidence; the grader decides."

    head = ["Feature", "Max", "Proposed", "Evidence for the proposed mark", "Where to check it"]
    for i, h in enumerate(head, start=1):
        c = ws.cell(row=5, column=i, value=h)
        c.font, c.fill, c.border, c.alignment = hdr, fill, thin, wrap

    r = 6
    for name, mx, prop, why, where in ROWS:
        ws.cell(row=r, column=1, value=name).border = thin
        ws.cell(row=r, column=2, value=mx).border = thin
        pc = ws.cell(row=r, column=3, value=prop); pc.border = thin; pc.font = Font(bold=True)
        for col, val in ((4, why), (5, where)):
            c = ws.cell(row=r, column=col, value=val); c.border = thin; c.alignment = wrap
        r += 1

    # Totals as live formulas, so the sheet cannot disagree with its own rows.
    ws.cell(row=r, column=1, value="TotalGrade").font = Font(bold=True)
    ws.cell(row=r, column=2, value=f"=SUM(B6:B{r-1})").font = Font(bold=True)
    ws.cell(row=r, column=3, value=f"=SUM(C6:C{r-1})").font = Font(bold=True)
    ws.cell(row=r, column=4,
            value="3D Game is the only row proposing 0. 110 of 115 available.").alignment = wrap

    for col, w in (("A", 30), ("B", 7), ("C", 10), ("D", 78), ("E", 46)):
        ws.column_dimensions[col].width = w
    for row in range(6, r):
        ws.row_dimensions[row].height = 76
    ws.freeze_panes = "A6"

    wb.save(out_path)
    total = sum(x[2] for x in ROWS)
    mx = sum(x[1] for x in ROWS)
    print(f"wrote {out_path}  proposed {total} / {mx}")

if __name__ == "__main__":
    build(Path(__file__).resolve().parent.parent / "submission_documents" / "Group52_Rubric_SelfAssessment.xlsx")

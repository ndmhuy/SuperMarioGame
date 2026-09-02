#!/usr/bin/env python3
"""build_contributions.py -- generate Member_Contributions.md and .xlsx from one source.

Why this exists: before this script, the two submission editions drifted by
construction. Member_Contributions.xlsx was written directly with openpyxl on
2026-08-31, and Member_Contributions.md was typed by hand to match it -- two
copies of the same facts, kept in sync only by a human remembering to update
both (exactly the failure g-rule-22 names: "Reference Docs Are Derived, Never
Hand-Synced"). This script is the fix: TASKS below is the one place task rows,
descriptions and hour estimates live; MEMBERS is the one place student
identity and the manual Percent convention live; git commit counts are
measured live from `git shortlog -sn` rather than typed by hand. Both editions
are rendered from these in one run, so they cannot drift apart again.

Usage:
    python3 scripts/build_contributions.py

Requires: openpyxl (3.1.5 verified working).
"""
from __future__ import annotations

import subprocess
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REPO_ROOT = Path(__file__).resolve().parent.parent
MD_PATH = REPO_ROOT / "submission_documents" / "Member_Contributions.md"
XLSX_PATH = REPO_ROOT / "submission_documents" / "Member_Contributions.xlsx"

COURSE = "CS202 - C++ Programming / Object-Oriented Design"
CLASS_NAME = "25A01"
GROUP = "52"
PROJECT_SCORE = 10

# --- Member roster -----------------------------------------------------
# `percent` is the manual PA-weighting convention (see the honesty note this
# script generates below) -- it is a judgment call the raw Tasks/Hours/Git
# percentages *inform*, not a number derived from them. Keep it in sync with
# what the group actually agrees to file; nothing here computes it for you.
MEMBERS = [
    {"student_id": "25125083", "full_name": "Nguyễn Đình Minh Huy", "percent": 0.5},
    {"student_id": "25125084", "full_name": "Trần Gia Huy", "percent": 0.5},
]

# --- Git identity resolution --------------------------------------------
# Maps a `git shortlog` author name to a student_id above. Per AGENTS.md's
# Layer 2 "Member Identity Resolution": FubuGold is Trần Gia Huy's second
# git identity (a GitHub noreply commit address), i.e. Member B.
# An author name that shows up in `git shortlog -sn` but is not listed here
# makes the script fail loudly (see count_git_commits) rather than silently
# under-count someone's work.
GIT_AUTHOR_TO_STUDENT_ID = {
    "Nguyen Dinh Minh Huy": "25125083",
    "Tran Gia Huy": "25125084",
    "FubuGold": "25125084",
}

# --- Task table -----------------------------------------------------------
# One row per unit of logged work: (student_id, task_description, hours, evidence).
# Rows 1-120 (unchanged) came from the group's weekly reports and the git/log
# evidence cited inline. Rows 121+ extend the table with the 2026-08-31 ..
# 2026-09-02 sessions, one row per logs/agent_history.log entry's headline,
# grouped where several entries cover one continuous unit of work; hours are
# estimated the same way rows 106-120 were (session scope, files touched,
# summary density) -- they are estimates, not measurements, same as every
# other hour in this table.
TASKS: list[tuple[str, str, float, str]] = [
# --- original 120 rows, carried over from the hand-written table (unchanged) ---
    ("25125083", "Configured CMake (FetchContent) for SFML 3.0.2 + Dear ImGui/ImGui-SFML and set up the project directory structure", 2, "see docs/Group52_04/52.md"),
    ("25125083", "Drafted the v2.0 project spec (110 features), UML class diagrams and the implementation plan (TASKS.md, TASK_DIVISION.md)", 3, "see docs/Group52_04/52.md"),
    ("25125083", "Wrote clean C++ base skeletons for managers, physics and entities", 1.5, "see docs/Group52_04/52.md"),
    ("25125083", "Implemented MathUtils (vector ops, clamp, interpolation helpers)", 1, "see docs/Group52_04/52.md"),
    ("25125083", "Built GameStateManager/IGameState state-stack pattern and the Game singleton main loop at a fixed 60Hz timestep", 3, "see docs/Group52_04/52.md"),
    ("25125083", "Implemented AABB collision primitives and a dynamic SpatialHash broadphase grid", 2.5, "see docs/Group52_04/52.md"),
    ("25125083", "Built the PhysicsEngine update pipeline (gravity, gravity zones/water viscosity, axis-separated tile resolution)", 2.5, "see docs/Group52_04/52.md"),
    ("25125083", "Implemented CollisionDetector/CollisionResolver (solid tiles, conveyors, stomps/damage, player-vs-player bounces)", 3, "see docs/Group52_04/52.md"),
    ("25125083", "Restructured Entity attributes to protected with friend-class access and read-only getters", 1, "see docs/Group52_04/52.md"),
    ("25125083", "Implemented the IPlayerState decorator pattern (StarDecorator, MegaDecorator) with timers", 2, "see docs/Group52_04/52.md"),
    ("25125083", "Coded concrete playable characters (Mario, Luigi, Toad, Peach) with custom physics/abilities", 2.5, "see docs/Group52_04/52.md"),
    ("25125083", "Built the Item base class and 12 concrete item subclasses (Mushroom, FireFlower, Coin, Star, etc.)", 3, "see docs/Group52_04/52.md"),
    ("25125084", "Built InputManager (Command Pattern) with dual-player key mapping (P1 WASD/Space/F, P2 Arrows/M)", 2, "see docs/Group52_04/52.md"),
    ("25125084", "Implemented SoundManager (Singleton) with music/SFX channels, WAV/OGG loading", 1.5, "see docs/Group52_04/52.md"),
    ("25125084", "Researched/designed the IMovementStrategy interface and 8 enemy AI strategies", 2, "see docs/Group52_04/52.md"),
    ("25125084", "Prepared abstract skeletons for the Enemy and Block type hierarchies", 1, "see docs/Group52_04/52.md"),
    ("25125084", "Designed the EntityFactory registry blueprint for JSON-driven instantiation", 1, "see docs/Group52_04/52.md"),
    ("25125083", "Implemented the full Camera class (LERP tracking, screen shake, viewport clamping)", 2, "see docs/Group52_05/52.md"),
    ("25125083", "Refactored move/run commands to intent flags on Character/Player instead of direct velocity changes", 1.5, "see docs/Group52_05/52.md"),
    ("25125083", "Centralized friction/acceleration/run-speed logic into PhysicsEngine::update() with surface-specific properties", 2, "see docs/Group52_05/52.md"),
    ("25125083", "Integrated Camera into PlayingState with dual-view (world vs HUD) rendering", 1.5, "see docs/Group52_05/52.md"),
    ("25125083", "Added mid-air wall-slide friction with a capped slide speed (WALL_SLIDE_SPEED)", 1.5, "see docs/Group52_05/52.md"),
    ("25125083", "Integrated nlohmann/json and built LevelLoader to parse tile/entity/spawn data from level JSON", 2.5, "see docs/Group52_05/52.md"),
    ("25125083", "Authored 4 level files (level_1..3, bonus_1) and an ImGui level selector for playtesting", 2, "see docs/Group52_05/52.md"),
    ("25125084", "Implemented IMovementStrategy and 8 concrete AI strategies (Patrol, Chase, Fly, etc.)", 2.5, "see docs/Group52_05/52.md"),
    ("25125084", "Coded the Enemy base class plus Goomba, KoopaTroopa, KoopaParatroopa, Boo", 3, "see docs/Group52_05/52.md"),
    ("25125084", "Coded the Block base class plus BrickBlock, QuestionBlock, Pipe, Flagpole", 2.5, "see docs/Group52_05/52.md"),
    ("25125084", "Coded 7 advanced v2.0 enemies (PiranhaPlant, BulletBill, HammerBro, Thwomp, ChainChomp, Lakitu, Spiny)", 3.5, "see docs/Group52_05/52.md"),
    ("25125084", "Coded 5 advanced v2.0 blocks (HiddenBlock, MovingPlatform, FallingPlatform, IceBlock, ConveyorBelt)", 2.5, "see docs/Group52_05/52.md"),
    ("25125084", "Implemented EntityFactory to instantiate 25+ entity types from JSON", 2.5, "see docs/Group52_05/52.md"),
    ("25125084", "Built SpriteSheet, Animation (flyweight) and AnimationManager/Animator", 3, "see docs/Group52_05/52.md"),
    ("25125084", "Wrote verify_enemies_new.cpp / verify_blocks_new.cpp test harnesses", 1.5, "see docs/Group52_05/52.md"),
    ("25125083", "Built Serializer for save-slot JSON schemas (character, player state, scores, checkpoints, playtime)", 2.5, "see docs/Group52_06/52.md"),
    ("25125083", "Implemented StatisticsTracker (EventBus subscriber for coins/enemies/deaths/combos/playtime)", 1.5, "see docs/Group52_06/52.md"),
    ("25125083", "Implemented AchievementManager (10 achievements) plus a slide-in/out toast HUD", 2, "see docs/Group52_06/52.md"),
    ("25125083", "Built the ImGui Pause/Settings panel (save slots, volume, keybindings, difficulty, auto-save)", 2, "see docs/Group52_06/52.md"),
    ("25125083", "Wrote the verify_save_load.cpp integration test suite", 1, "see docs/Group52_06/52.md"),
    ("25125083", "Implemented the Undo/Redo command framework (EditorCommands)", 2, "see docs/Group52_06/52.md"),
    ("25125083", "Built the MapEditor viewport dashboard (tile/entity brushes, undo/redo, save)", 2, "see docs/Group52_06/52.md"),
    ("25125083", "Implemented pixel-to-grid coordinate mapping and grid-line rendering", 1.5, "see docs/Group52_06/52.md"),
    ("25125083", "Centralized duplicate tile/entity name mappers into SerializationUtils", 1, "see docs/Group52_06/52.md"),
    ("25125083", "Wrote the verify_map_editor.cpp test harness", 1, "see docs/Group52_06/52.md"),
    ("25125084", "Participated in architectural reviews, resolved cross-branch merge conflicts, refined achievement SFX triggers", 1, "see docs/Group52_06/52.md"),
    ("25125083", "Merged A/feature/save-load and A/feature/level-editor; addressed PR review fixes (EventBus sentinel, debug key remap, Game keybinding encapsulation)", 2, "see docs/Group52_07/52.md"),
    ("25125083", "Fixed a stale achievement-toast bug on slot reload; aligned achievement keys with SPEC.md", 1, "see docs/Group52_07/52.md"),
    ("25125083", "Built the consolidated verify_all.cpp test runner combining 6 test suites; updated CMakeLists.txt", 2, "see docs/Group52_07/52.md"),
    ("25125083", "Authored two_player_ai_plan.md (2-player + Shadow Mario architecture design doc)", 2.5, "see docs/Group52_07/52.md"),
    ("25125083", "Ran an interactive binary demo to validate save/load and level-editor integration", 1, "see docs/Group52_07/52.md"),
    ("25125084", "Entity Factory & concrete entities merged into dev (integration of prior work)", 0.5, "see docs/Group52_07/52.md"),
    ("25125084", "Enemy hierarchy + AI strategies merged into dev (integration of prior work)", 0.5, "see docs/Group52_07/52.md"),
    ("25125084", "Blocks/platforms merged into dev (integration of prior work)", 0.5, "see docs/Group52_07/52.md"),
    ("25125084", "Built Hud (lives/score/coins/time/stage) plus a boss health bar and retro typography (PressStart2P/SuperMario256 fonts)", 2, "see docs/Group52_07/52.md"),
    ("25125084", "Finalized SpriteSheet/AnimationManager/Animator; added verify_blocks/verify_enemies/verify_graphics/verify_graphics_visual/verify_hud_boss test targets", 1.5, "see docs/Group52_07/52.md"),
    ("25125084", "Consolidated duplicate asset/ and assets/ folders into one assets/ root", 0.5, "see docs/Group52_07/52.md"),
    ("25125084", "Built the Minimap component (scaled tilemap preview, camera frustum, player tracking dot)", 2, "Git commit cb6eab2"),
    ("25125084", "Wrote the verify_minimap_visual.cpp test target", 1, "see docs/Group52_07/52.md"),
    ("25125083", "Reviewed/validated the ParticleSystem/ObjectPool<Particle> design (O(1) acquire/release, zero-allocation)", 1, "see docs/Group52_08/52.md"),
    ("25125083", "Updated TASKS.md through Phase 8 and Bonus A, cross-referenced against TASK_DIVISION.md", 0.5, "see docs/Group52_08/52.md"),
    ("25125083", "Ran git fetch --all, verified dev in sync with origin/dev, tracked branch status", 0.5, "see docs/Group52_08/52.md"),
    ("25125084", "Implemented ParticleSystem with ObjectPool<Particle> (burst/continuous emission modes)", 2.5, "Git commit 7a6f62b"),
    ("25125084", "Tuned particle parameters (alpha fade, emission rate, spread angle)", 1.5, "Git commits c7d4fd4, 62c78ef"),
    ("25125084", "Built the verify_particles_visual interactive test target", 1, "see docs/Group52_08/52.md"),
    ("25125084", "Extracted the player sprite sheet player.json (92 animation frames: idle/walk/run/jump/etc.)", 3, "Git commits d6bd07d, 0be4508, 9403be5"),
    ("25125084", "Built the verify_player_animation_visual test target", 1, "see docs/Group52_08/52.md"),
    ("25125084", "Began sprite naming standardization / asset pipeline scripts", 1, "Git commit 15327b4"),
    ("25125084", "Enhanced random-number generation with std::uniform_real_distribution for particle spread", 0.5, "Git commit 9bb9d16"),
    ("25125083", "Implemented the Fireball entity + ObjectPool<Fireball>, EntityFactory/EventBus wiring, verify_fireball.cpp (4/4 pass)", 3, "see docs/Group52_09/52.md"),
    ("25125083", "Enhanced Flagpole with height-scaled scoring, animation and a LevelComplete event; verify_flagpole_trampoline.cpp", 2, "see docs/Group52_09/52.md"),
    ("25125083", "Enhanced Trampoline bounce mechanics (spring compression, configurable impulse)", 1.5, "see docs/Group52_09/52.md"),
    ("25125083", "Built the CheckpointFlag entity (visual states, auto-save, CheckpointActivated event, respawn coordinates)", 2, "see docs/Group52_09/52.md"),
    ("25125083", "Implemented the Memento pattern: GameSnapshot + TimeRewindManager (300-frame circular buffer); verify_memento_rewind.cpp (4/4 pass)", 3.5, "see docs/Group52_09/52.md"),
    ("25125083", "Built MapGenerator for procedural level generation; verify_map_generator.cpp", 3, "see docs/Group52_09/52.md"),
    ("25125083", "Implemented the dual bounding-box architecture (physicsBox/combatBox) across Entity/CollisionDetector/CollisionResolver", 2.5, "see docs/Group52_09/52.md"),
    ("25125083", "Added void-fall death detection in PlayingState", 1, "see docs/Group52_09/52.md"),
    ("25125083", "Added brick/question-block head-butt reactions in PhysicsEngine", 1.5, "see docs/Group52_09/52.md"),
    ("25125083", "Removed public setLives/setCoins/setScore from Player; added friend class PlayingState", 0.5, "see docs/Group52_09/52.md"),
    ("25125083", "Added Camera::move() for editor WASD/arrow panning", 1, "see docs/Group52_09/52.md"),
    ("25125083", "Added a Level Editor Mode (F1) button to MenuState", 0.5, "see docs/Group52_09/52.md"),
    ("25125083", "Hardened Game::shutdown() to pop all game states before manager teardown", 1, "see docs/Group52_09/52.md"),
    ("25125083", "Added multi-path font-loading fallback candidates in PlayingState", 0.5, "see docs/Group52_09/52.md"),
    ("25125083", "Fetched/merged origin/B/feature/graphic-visual into dev; resolved font-loading + agent_history.log conflicts; pushed dev", 1, "Git commit 17cef23"),
    ("25125083", "Fixed non-solid TileType::Coin tiles not being collected on walk-through", 1, "Git commit d12e763"),
    ("25125083", "Deduplicated a ResourceManager missing-SoundBuffer warning spam", 0.5, "see docs/Group52_09/52.md"),
    ("25125084", "Built a spritesheet-merging tool; updated enemy_projectile/item/world_scenery_item atlases", 3, "Git commits 044a5d8, 80c048e, b90a899"),
    ("25125084", "Cleaned sprite/atlas coordinates affected by the merge", 1, "Git commit 10828a7"),
    ("25125084", "Updated SPRITE_NAMING_CONVENTIONS.md (frame inventories, naming standards)", 1, "see docs/Group52_09/52.md"),
    ("25125084", "Built SpriteColorFilter (HSL->RGB, 12Hz Star Power cycling, 15Hz hurt flicker, hit-flash tint)", 2, "Git commit 2c3bb83"),
    ("25125084", "Built SpriteTransformAnim (scale-lerp easing, 720 deg/s spin, 12Hz Mega/Mini flicker)", 2, "Git commit 2c3bb83"),
    ("25125084", "Built the EntityDeathEffect singleton (EnemyFlip, StarKillSpin, PlayerDeathHop)", 2, "Git commit 2c3bb83"),
    ("25125084", "Added a SoundManager synthesized fallback beep buffer", 1, "see docs/Group52_09/52.md"),
    ("25125084", "Upgraded verify_graphics_visual with a Transform/Color-Filter FX ImGui tab", 1.5, "see docs/Group52_09/52.md"),
    ("25125083", "Refined MapGenerator (elevation profiles, biome ceilings, lava pits, jump-reachability-guarded platforms, 3 Star Coins, threat-pacing curve) + ImGui tuning panel", 3.5, "see docs/Group52_10/52.md"),
    ("25125083", "Built the 3-world campaign + warp-pipe system; generate_game_levels.cpp synthesizing 7 level JSONs; extended SerializationUtils/LevelLoader for warp metadata; seamless sub-level transitions", 4, "see docs/Group52_10/52.md"),
    ("25125083", "Added ImGui campaign navigation dropdowns and an Active Level Tube warp selector", 1.5, "see docs/Group52_10/52.md"),
    ("25125083", "Fixed a static-block gravity bug (getGravityMultiplier() override); added item resting-state physics (resolveItemVsBlock)", 1.5, "see docs/Group52_10/52.md"),
    ("25125083", "Optimized player hitbox dimensions (32px to 24px width tiers); refined stomp-vs-side-hit combat resolution", 2, "see docs/Group52_10/52.md"),
    ("25125083", "Fixed the LevelLoader ifstream failbit cascade (file.clear()); expanded fallback search paths", 1.5, "see docs/Group52_10/52.md"),
    ("25125083", "Fixed a SIGABRT shutdown crash (explicit m_window.close()) and hardened GameStateManager transitions", 1.5, "see docs/Group52_10/52.md"),
    ("25125084", "Wired sprite animations across all entities; aspect-ratio scaling for HUD/items; POW 8-frame shockwave; trampoline/PSwitch visual states", 3.5, "Git commits 069817b, 0bc21c3, 2d5058e, 83660b5, 065557d, 61e30bc"),
    ("25125084", "Built the Web Metadata Editor Studio (metadata_editor.html, edit_sprite_metadata.py) with cursor-anchored zoom, position-finder HUD, drag-to-draw frames, TexturePacker-compatible export", 2.5, "see docs/Group52_10/52.md"),
    ("25125084", "Integrated sprite HUD assets (animated coin counter, Star Coin indicators, character-specific lives icon)", 1.5, "see docs/Group52_10/52.md"),
    ("25125084", "Built PipeRenderer (2-block-wide AABB, directional multi-part pipe head/body rendering)", 1.5, "Git commit a0f1a3b"),
    ("25125084", "Wired game-wide SFX/BGM into SoundManager (jump, stomp, coin, break, power-up/down, boing, stage-clear, game-over) + fallback buffer", 2.5, "Git commits 1002595, c383c19, 9976209"),
    ("25125084", "Tuned all 11 enemy AI state machines (Thwomp slam cycle, Boo gaze-detection, HammerBro parabolic throw, Koopa shell physics)", 3, "Git commit d954a66"),
    ("25125084", "Built verify_enemies_behavior.cpp + enhanced verify_enemies_visual.cpp (11 isolated test rooms)", 1.5, "Git commit d954a66"),
    ("25125083", "Resolved the Level 1-3 crash on Bowser's first fireball (mid-frame entity-spawn vector reallocation); rebuilt and playtested on Windows", 3, "see logs/agent_history.log 2026-08-16/22"),
    ("25125083", "Fixed 12 defects from a playtest pass plus two menus that overflowed", 2, "Git commit f6fc8bc"),
    ("25125083", "Fixed CMakeLists naming src/Graphics/HUD.cpp when the file is Hud.cpp", 0.5, "Git commit 03bf828"),
    ("25125083", "Fixed the player being unable to jump while walking into a wall", 1, "Git commit 60cc735"),
    ("25125083", "Fixed 28 files using std:: symbols without including their header", 1, "Git commit 4ac8a1c"),
    ("25125083", "Fixed CI building only 2 of the 13 binaries it then tested", 1, "Git commit 8141120"),
    ("25125083", "Fixed 4 test harnesses opening an sf::RenderWindow on a CI runner with no display", 1, "Git commit 33d5aae"),
    ("25125083", "Fixed three defects the screenshots showed that the tests did not (backdrop ground line, hill-bottom offset, a Player 2 HUD badge overwriting the world field)", 2, "Git commit da8ca99"),
    ("25125083", "Sealed test hermeticity: Serializer::setSaveDirectory() + TestSaveSandbox.hpp so all 23 verify_* harnesses point at a fresh temp directory", 2.5, "Git commit 045f9b7"),
    ("25125083", "Generated real UML class diagrams from the headers (tools/gen_class_diagram.py) and wrote the CS202 final report's first edition (15 sections, generated by build_report.py)", 4, "Git commits a0395d6, 3c0a768, 64088b4"),
    ("25125083", "Produced a LaTeX/PDF edition of the report in the HCMUS template and deepened it (implementation section 3->11 subsections, technical-problems section rebuilt into 8 themed groups)", 4, "Git commits fdec9cc, 730336c"),
    ("25125083", "Diagnosed and fixed three defects reported from a real playthrough: the end-of-level castle's buffer margin, Boom Boom's arena overlapping the 1-2 flagpole, and the camera/z-order bug carried into 1-3; added regression tests for all three", 4, "this session, branch A/fix/level-completion-and-camera-defects"),
    ("25125083", "Implemented Endless Mode (chunked infinite level generation with rising difficulty, distance-based scoring) and a real solvability oracle for MapGenerator, after evaluating (and rejecting the framework of) the abandoned A/mapgen-gan-plan and A/rl-neural-policy side branches for anything worth porting", 5, "this session, branch A/feature/endless-mode-and-solvability"),
    ("25125083", "Rebuilt the submission package: a 97-feature list and per-member manual test checklist grounded in a fresh source-code audit, the Member Contributions spreadsheet (matching the DesignPatternsGroup52 format), and an expanded report (OOP/pattern rationale, two new UML diagrams)", 4, "this session, branch A/docs/submission-package-update"),
    ("25125084", "Resolved merge conflicts while pulling the latest changes and merged into main", 2, "Git commits 72b13c6, a522007, 75e9cbb"),

# --- new rows: 2026-08-31 .. 2026-09-02 sessions, one row per log entry's
# headline, grouped where multiple entries cover one unit of work (see
# logs/agent_history.log for the full detail behind each summary) ---
    ("25125083", "Merged all outstanding work to dev/main and generated the initial submission package (AI Usage Declaration, features list, task division, demo-video placeholder, member contributions, final report) while kicking off the defect-auditing test suite.", 3, "see logs/agent_history.log 2026-08-30, branch main"),
    ("25125083", "Merged the level-completion, endless-mode, and submission-package branches into dev then main and pushed; corrected the AI Usage Declaration's tool attribution (Antigravity vs. Claude Code) per user feedback.", 1, "see logs/agent_history.log 2026-08-31 09:45 and 10:00, branch main"),
    ("25125083", "Replaced the report's ASCII architecture diagram with a real figure, embedded full-depth UML diagrams in the appendix, and rewrote the class-diagram generator to render every attribute/method with UML visibility notation, fixing a brace-init parsing bug along the way.", 3.5, "branches A/docs/report-architecture-figure-and-full-uml, A/docs/detailed-uml-full-attributes-methods"),
    ("25125083", "Ran a three-way SPEC/code/report audit via three parallel subagent passes, producing a 15-defect, 14-task (R1-R14) remediation plan and correcting nine stale claims in the features list and report.", 4, "branch A/docs/spec-feature-audit"),
    ("25125083", "Small-defect and hygiene batch (R1/R2/R3/R6): fixed SoundManager's level-BGM index mapping, Enemy/Spiny's screen-height despawn bug, and unsurfaced solvability results; deleted the dead AnimationManager class; degraded audio startup gracefully with no device; untracked .member_profile.json and added SPEC's descope addendum.", 3.5, "see logs/agent_history.log 2026-08-31 11:48-12:34"),
    ("25125083", "Migrated every raw EventBus SubscriptionId to ScopedSubscription across PlayingState/AchievementManager/StatisticsTracker/Camera (R4), and fixed a real static-destruction-order SIGSEGV the migration surfaced.", 2.5, "branch A/fix/scoped-subscription-adoption, commit 00c08f8"),
    ("25125083", "Fixed a P-Switch soft-lock caused by mis-themed sub-level floors, a half-rendered single-column pipe in all three sub-levels, and made the overworld QuestionBlock's sprite reflect its emptied state to match SPEC (R16).", 3, "branch A/fix/sublevel-tile-batch, commit cc6a32d"),
    ("25125083", "Investigated reported 'Bowser never spawns' and 'flag reachable mid-fight' defects (both refuted as already fixed/non-reproducing) and widened Boom Boom's arena from 3.5 to 12.5 tiles of pacing room by removing decorative staircase tiles (R15).", 2.5, "branch A/fix/boss-encounter-batch, commit 56c8db7"),
    ("25125083", "Fixed the end-game audio batch: castle_complete fanfare being clobbered by level_complete on boss levels via a deferred one-shot-music swap; investigated and struck two further audio defects as not reproducible/already fixed (R19).", 2, "branch A/fix/endgame-audio-batch, commit de52042"),
    ("25125083", "Reconciled the SPEC-audit document's defect ledger against shipped code, marking or striking about twenty defects and phases (R1-R19) and documenting four previously mis-recorded playtest defects.", 1, "branch A/docs/audit-status-reconciliation, commit cc08fc3"),
    ("25125083", "Fixed the Star power's missing rainbow-tint visual, the sub-level Piranha Plant's pipe-centering offset, and the fireball-kill flip-death sprite anchor via a shared Entity::drawSprite fix reaching every flip-killable enemy (R17).", 2.5, "branch A/fix/player-enemy-visual-batch, commit 4550bfe"),
    ("25125083", "Eliminated all 28 real dynamic_casts from CollisionResolver/PhysicsEngine via new virtual Entity/Item/Block/Enemy/Character hooks (nine commits, one dispatch family each), adding 26 regression checks and discovering that Release's -DNDEBUG compiled away nearly all existing assert()-based coverage (R5).", 4, "branch A/refactor/collision-dispatch, commit 4b87e0d"),
    ("25125083", "Placed all six previously-unused enemy types (KoopaParatroopa, Boo, BulletBill, Thwomp, ChainChomp, Lakitu) and hidden blocks across the campaign, fixed the secret_finder achievement's wrong trigger condition, and fixed two playtest-found defects in Bullet Bill and Lakitu placement (R9).", 4, "branch A/feature/campaign-population-pass, commit 4d6be7e"),
    ("25125083", "Stripped -DNDEBUG from ctest targets only (fixing vacuously-passing assert-based harnesses), refactored the build to compile app sources once via two CMake OBJECT libraries, and added a mutation-tested guard against tests writing into the real saves/ directory (R11).", 4, "branches A/fix/tests-honour-assertions, A/build/object-library, A/fix/hermetic-tests"),
    ("25125083", "Wired eight dormant systems into real gameplay: star-power side-touch kills, four unused particle types, surface-dependent footsteps, combo-hit pitch escalation, menu-row navigation SFX, a single-player camera clamp, and debug-console tab-completion (R7).", 3, "branch A/feature/wire-dormant-vfx-sfx, commit 7d94fb9"),
    ("25125083", "Added a keyboard-navigable LOAD GAME menu with a 3-slot picker showing character/level/score/time previews, reusing PlayingState's existing loadFromSlot path (R8).", 2.5, "branch A/feature/load-game-menu, commit d6672b7"),
    ("25125083", "Implemented main-menu Attract Mode (30s idle triggers a bundled demo replay, dismissed by any key), fixing two latent ReplayRecorder playback-pacing/fade-in bugs surfaced by actually watching it run (R10).", 3, "branch A/feature/attract-mode, commit 6075c27"),
    ("25125083", "Ran a full scripted playtest evidence pass across campaign completion, P-Switch/POW/axe, 2P versus death, difficulty modes, key rebinding, and FPS, discovering two new defects: boss destroyed by the respawn-safety sweep, and Load Game always resuming World 1-1 (R12).", 3, "branch A/verify/full-playtest-pass"),
    ("25125083", "Fixed the boss-survives-respawn defect (D29), Load Game restoring the saved level (D30), and re-implemented a double-exit guard for PlayingState, each with a mutation-tested regression case; merged and independently re-verified into dev (R20).", 3, "branch A/fix/remaining-defects-d29-d30-exit, commit 6e4868d"),
    ("25125083", "Root-caused and fixed a user-reported bug where Bowser fell out of the world through his own bridge/lava floor, since the void-plane check only covered players; added an Entity::onLeftLevel virtual hook.", 2, "branch A/fix/boss-falls-out-of-world, commit 95521a8"),
    ("25125083", "Fixed ten gameplay/visual issues in one pass: castle sprite/position, level-clear music timing, Star invincibility flicker, Piranha Plant centering, fireball flip pivot, Bowser arena verification, bonus-level plateau/castle seating, and Peach's Space-key float.", 3, "see logs/agent_history.log 2026-09-01 15:31 and 16:52, branch dev"),
    ("25125083", "Fixed sub-level pipe teleports and half-sprite rendering, decoupled the S key from ground-pound (crouch-only), halved Thwomp slam speed with a narrowed detection column, rewrote Chain Chomp's tethered-chase AI, enclosed the 1-3 bridge with walls, and fixed the camera snap on Bowser's defeat.", 3, "see logs/agent_history.log 2026-09-01 17:55, branch dev"),
    ("25125083", "Wrote the demo-video-requirements checklist; fixed the M-key minimap toggle, a duplicate player-death animation, metal-footstep SFX volume, and further tuned Thwomp slam/recovery speed and its 1-3 placement.", 1.5, "see logs/agent_history.log 2026-09-01 23:30, branch dev"),
    ("25125083", "R21 Phase 1 defect batch: fixed twelve release defects in one pass — a duplicate untracked asset tree causing level-data drift, half/warped pipe rendering, spawns embedded in solid tiles, an incomplete castle slab, double-integrated moving platforms, menu text overflow, duplicate achievement toasts, and gated debug ImGui off by default.", 4, "branch A/release/defect-batch-r21, commit 171a5b3"),
    ("25125083", "Merged the R21 batch and completed Phases 2-4: built a real GUI level editor wired to EntityFactory (fixing 24 of 40 silently-broken palette buttons), debug/recording cheats with a void-rescuing Immortal mode, fixed far-chunk procedural entities teleporting back to chunk-local coordinates, an explicit entity-type registry, and Bonus D's dynamic day/night lighting via a real GLSL shader.", 4.5, "commit 97fc43a (merge into dev), branch A/release/defect-batch-r21"),
    ("25125083", "Second user-reported defect wave: fixed warp-pipe hitbox/art and entry-mode rendering, Bowser's fireball stagger triple-counting a single fireball as three hits, Bowser falling into lava, Lakitu dropping a Fire Flower to keep the fight winnable, the 2P camera freezing on the eliminated player instead of following the survivor, and added save-slot choice/delete.", 3.5, "commit e97b7da (merge), branch A/fix/versus-camera-and-axes"),
    ("25125083", "Audited OOP/SOLID adherence and design-pattern fidelity across the codebase (dynamic_cast counts, singleton count, friend-class audit, PDF text-overlap root cause) and authored the phased submission-sweep plan with per-lane dependencies, models, and effort levels.", 3, "docs/issues/submission_sweep_plan_2026-09-02.md, branch dev"),
    ("25125083", "Ran the submission-sweep Phase 0 baseline: clean rebuild, full ctest/regression run, a live smoke-test launch, and confirmed the LevelSolvability fix, with no code changes.", 1, "see logs/agent_history.log 2026-09-02 18:58, branch dev"),
    ("25125083", "Wrote the W11, W12, and W13 weekly reports from git log and agent-history evidence, including tracing the World 1-3 mid-frame entity-spawn crash's root cause for W11.", 3, "branches A/docs/weekly-w11, A/docs/weekly-w12-w13"),
    ("25125083", "Archived 20+ stale root/docs artifacts to docs/archive/, untracked stale saves/PDFs/zips, then reverted an AGENTS.md edit that had landed inside AgentHub's generated block and routed the rule change upstream instead.", 2, "branch A/chore/archive-2026-09-02, commit 9b83c9a"),
    ("25125083", "Extended the AI Usage Declaration to cover the full R1-R21 remediation plan, then corrected its date range and removed an unverifiable test-count claim per review.", 1, "branch A/docs/ai-declaration, commit 96560bd"),
    ("25125083", "Captured fresh gameplay screenshots and composited three YouTube thumbnail candidates, then narrowed to one final crop-and-zoom version after review found Mario illegibly small at native resolution.", 1.5, "branch A/docs/video-thumbnail"),
    ("25125083", "Fixed the report PDF's overlapping-table-text defect with content-aware column sizing, added a build.sh Overfull-hbox guard, and split two illegible full-detail UML pages into landscape multi-page spreads.", 2.5, "branch A/docs/report-layout, commit 0f3e5ea"),
    ("25125083", "Fixed the level editor's F5 playtest routing to Game Over/main menu instead of popping back to the editor on death, with a mutation-tested regression case.", 1.5, "branch A/fix/editor-playtest-gameover, commit 2bd2f80"),
    ("25125083", "Fixed MapGenerator's procedurally-generated vault exit pipe defaulting to the wrong entry mode; investigated the reported 'unreachable third Hard-mode axe' and found it does not reproduce, shipping a protective regression guard instead of an unjustified data edit.", 1.5, "branch A/fix/hard-axe-and-vault-pipe"),
    ("25125083", "Added a camera-distance gate on Endless Mode's ever-growing entity list, added ImGui sliders for the dynamic-lighting tunables, and deleted the dead UiRenderer::wrapText function.", 2, "branch A/fix/endless-gate-lighting-sliders-wraptext, commit 0fe7138"),
    ("25125083", "Fixed MovingPlatform never advancing its own animation, flipped Spiny's default to spawn as an unhatched egg matching SPEC, and added a hardened post-condition regression test after review found the original test too weak.", 2, "branch A/fix/moving-platform-and-spiny-egg, commit cf195cb"),
]


def count_git_commits(repo_root: Path) -> dict[str, int]:
    """Return {student_id: commit_count} measured live from `git shortlog -sn`.

    Raises if an author appears in the repo history that isn't in
    GIT_AUTHOR_TO_STUDENT_ID -- an unmapped author would otherwise silently
    vanish from the commit count instead of failing loudly.
    """
    result = subprocess.run(
        ["git", "shortlog", "-sn", "HEAD"],
        cwd=repo_root,
        capture_output=True,
        text=True,
        check=True,
    )
    counts: dict[str, int] = {m["student_id"]: 0 for m in MEMBERS}
    for line in result.stdout.splitlines():
        line = line.strip()
        if not line:
            continue
        count_str, author = line.split("\t", 1) if "\t" in line else line.split(None, 1)
        author = author.strip()
        try:
            n = int(count_str.strip())
        except ValueError:
            continue
        if author not in GIT_AUTHOR_TO_STUDENT_ID:
            raise ValueError(
                f"git shortlog shows author {author!r} ({n} commits) with no "
                "entry in GIT_AUTHOR_TO_STUDENT_ID -- resolve their identity "
                "(AGENTS.md Layer 2 'Member Identity Resolution') and add a "
                "mapping before regenerating, rather than silently dropping "
                "their commits from the count."
            )
        sid = GIT_AUTHOR_TO_STUDENT_ID[author]
        counts[sid] = counts.get(sid, 0) + n
    return counts


def get_git_sha(repo_root: Path) -> str:
    """Return the short HEAD sha, the same way count_git_commits() shells out."""
    result = subprocess.run(
        ["git", "rev-parse", "--short", "HEAD"],
        cwd=repo_root,
        capture_output=True,
        text=True,
        check=True,
    )
    return result.stdout.strip()


def build_summary(tasks, git_counts):
    """Compute per-member and total stats from TASKS and measured git counts."""
    per_member = {}
    for m in MEMBERS:
        sid = m["student_id"]
        rows = [t for t in tasks if t[0] == sid]
        per_member[sid] = {
            "student_id": sid,
            "full_name": m["full_name"],
            "percent": m["percent"],
            "tasks": len(rows),
            "hours": round(sum(t[2] for t in rows), 4),
            "git_commits": git_counts.get(sid, 0),
        }

    total_tasks = sum(v["tasks"] for v in per_member.values())
    total_hours = round(sum(v["hours"] for v in per_member.values()), 4)
    total_git = sum(v["git_commits"] for v in per_member.values())

    for v in per_member.values():
        v["tasks_pct"] = v["tasks"] / total_tasks if total_tasks else 0.0
        v["hours_pct"] = v["hours"] / total_hours if total_hours else 0.0
        v["git_pct"] = v["git_commits"] / total_git if total_git else 0.0

    return {
        "members": [per_member[m["student_id"]] for m in MEMBERS],
        "total_tasks": total_tasks,
        "total_hours": total_hours,
        "total_git": total_git,
        "num_students": len(MEMBERS),
        "max_percent": max(m["percent"] for m in MEMBERS),
    }


def gap_note(summary) -> str:
    """Generate the honesty note on the Git% vs Tasks/Hours% gap from live numbers."""
    members = sorted(summary["members"], key=lambda v: -v["git_pct"])
    lead = members[0]
    tasks_pcts = "/".join(f"{v['tasks_pct'] * 100:.0f}" for v in members)
    return (
        f"Note on the gap between the Git % column and the Tasks/Hours % columns: "
        f"{lead['full_name']}'s git identity carries a large share of small "
        f"fix/doc/log commits ({lead['git_commits']} of {summary['total_git']}) "
        f"accumulated across the whole session history, while task count and "
        f"estimated hours — a better proxy for actual design/implementation "
        f"effort — split closer to {tasks_pcts}. Both are shown rather than "
        f"only the more flattering one."
    )


def render_markdown(summary, tasks, measured_at: str) -> str:
    lines = []
    lines.append("# Member Contributions — Group 52 (Super Mario Game)")
    lines.append("")
    lines.append(f"**Course:** {COURSE}  ")
    lines.append(f"**Class:** {CLASS_NAME}  ")
    lines.append(f"**Group:** {GROUP}  ")
    lines.append("")
    lines.append(
        "This document is the same data as `Member_Contributions.xlsx` (the "
        "TA-facing spreadsheet, in the format used by the group's other CS202 "
        "project) rendered as Markdown/PDF for readability — both are "
        "generated from `scripts/build_contributions.py` (one source, one "
        "task table) so they cannot drift apart."
    )
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append("| | |")
    lines.append("|---|---|")
    lines.append(f"| Number of students | {summary['num_students']} |")
    lines.append(f"| Number of tasks | {summary['total_tasks']} |")
    lines.append(f"| Number of task hours | {summary['total_hours']:g} |")
    lines.append(f"| Number of Git commits | {summary['total_git']} |")
    lines.append(f"| Max student percentage | {summary['max_percent']:g} |")
    lines.append(f"| Project score | {PROJECT_SCORE} |")
    lines.append("")
    lines.append(f"*{measured_at}*")
    lines.append("")
    percent_convention = "/".join(f"{v['percent']:g}" for v in summary["members"])
    lines.append(
        "> Do not edit the grey cells in the spreadsheet edition. TAs enter "
        "the score of all PAs in the project score and get the individual "
        "scores; students may enter an estimated project score to see how "
        "percentages affect it. `Percent` is set to "
        f"{percent_convention} for both "
        "members — the raw Tasks/Hours/Git percentages below inform that "
        "judgment rather than dictate it, the same convention the reference "
        "spreadsheet uses."
    )
    lines.append("")
    lines.append(
        "| No | Student ID | Full name | Tasks | Tasks % | Task Hours | Hours % "
        "| Git Commits | Git % | Percent | Score (Student) |"
    )
    lines.append("|---|---|---|---|---|---|---|---|---|---|---|")
    for i, v in enumerate(summary["members"], start=1):
        lines.append(
            f"| {i} | {v['student_id']} | {v['full_name']} | {v['tasks']} | "
            f"{v['tasks_pct'] * 100:.1f}% | {v['hours']:g} | {v['hours_pct'] * 100:.1f}% | "
            f"{v['git_commits']} | {v['git_pct'] * 100:.1f}% | {v['percent']:g} | "
            f"{PROJECT_SCORE} |"
        )
    lines.append("")
    lines.append(gap_note(summary))
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## Task-by-task breakdown")
    lines.append("")
    lines.append(
        "*A task should be done by only 1-2 students. Each member worked far "
        "more than the minimum 5 tasks the reference template asks for.*"
    )
    lines.append("")
    lines.append("| No | Student ID | Full name | Task Description | Hours | Evidence |")
    lines.append("|---|---|---|---|---|---|")
    names_by_id = {m["student_id"]: m["full_name"] for m in MEMBERS}
    for i, (sid, desc, hours, evidence) in enumerate(tasks, start=1):
        lines.append(
            f"| {i} | {sid} | {names_by_id[sid]} | {desc} | {hours:g} | {evidence} |"
        )
    lines.append("")
    return "\n".join(lines)


def render_xlsx(summary, tasks, measured_at: str) -> Workbook:
    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="00DDDDDD", end_color="00DDDDDD", fill_type="solid")

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws["B1"] = "Course"
    ws["C1"] = COURSE
    ws["B2"] = "Class"
    ws["C2"] = CLASS_NAME
    ws["B3"] = "Group"
    ws["C3"] = GROUP
    ws["A4"] = "PROJECT CONTRIBUTION"
    ws["A4"].font = bold

    info_rows = [
        ("Number of students working in the project", summary["num_students"]),
        ("Number of tasks", summary["total_tasks"]),
        ("Number of task hours", summary["total_hours"]),
        ("Number of Git commits", summary["total_git"]),
        ("Max student percentage", summary["max_percent"]),
        ("Project score", PROJECT_SCORE),
    ]
    info_start = 5
    for offset, (label, value) in enumerate(info_rows):
        r = info_start + offset
        ws[f"C{r}"] = label
        ws[f"D{r}"] = value

    stamp_row = info_start + len(info_rows)
    ws[f"A{stamp_row}"] = measured_at

    instr1_row = stamp_row + 1
    instr2_row = instr1_row + 1
    ws[f"A{instr1_row}"] = (
        "Do not edit the grey cells. TAs will enter the score of all PAs in "
        "the project score and get the individual scores. Students can enter "
        "an estimated project score to see how percentages affects your scores"
    )
    ws[f"A{instr2_row}"] = (
        "Use columns Tasks - Percent, Task Hours - Percent, Git - Percent as "
        "references for column Percent"
    )

    header_row = instr2_row + 1
    headers = [
        "No", "Student ID", "Full name", "Tasks", "Tasks - Percent",
        "Task Hours", "Task Hours - Percent", "Git Commits", "Git - Percent",
        "Percent", "Score - Student", "Score - TA",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = bold
        c.fill = header_fill

    for i, v in enumerate(summary["members"], start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=v["student_id"])
        ws.cell(row=r, column=3, value=v["full_name"])
        ws.cell(row=r, column=4, value=v["tasks"])
        ws.cell(row=r, column=5, value=v["tasks_pct"])
        ws.cell(row=r, column=6, value=v["hours"])
        ws.cell(row=r, column=7, value=v["hours_pct"])
        ws.cell(row=r, column=8, value=v["git_commits"])
        ws.cell(row=r, column=9, value=v["git_pct"])
        ws.cell(row=r, column=10, value=v["percent"])
        ws.cell(row=r, column=11, value=PROJECT_SCORE)
        # Score - TA (column L) is left blank for the TA to fill in.

    col_widths_summary = {"A": 4, "B": 12, "C": 24, "D": 8, "E": 14, "F": 11,
                           "G": 18, "H": 12, "I": 12, "J": 9, "K": 14, "L": 12}
    for col, width in col_widths_summary.items():
        ws.column_dimensions[col].width = width

    # --- Tasks sheet ---
    ws2 = wb.create_sheet("Tasks")
    ws2["A1"] = (
        "A task should be done by only 1-2 students. Duplicate tasks are not "
        "recommended\nEach member should work on at least 5 tasks"
    )
    task_headers = [
        "No", "Student ID", "Full name", "Task Description", "Hours",
        "Evidence (screenshots of chat messages, of Git commits... proving "
        "that you perform and finish the task)",
    ]
    for col, h in enumerate(task_headers, start=1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = bold

    names_by_id = {m["student_id"]: m["full_name"] for m in MEMBERS}
    for i, (sid, desc, hours, evidence) in enumerate(tasks, start=1):
        r = 2 + i
        ws2.cell(row=r, column=1, value=i)
        ws2.cell(row=r, column=2, value=sid)
        ws2.cell(row=r, column=3, value=names_by_id[sid])
        ws2.cell(row=r, column=4, value=desc)
        ws2.cell(row=r, column=5, value=hours)
        ws2.cell(row=r, column=6, value=evidence)

    col_widths_tasks = {"A": 5, "B": 12, "C": 24, "D": 80, "E": 7, "F": 45}
    for col, width in col_widths_tasks.items():
        ws2.column_dimensions[col].width = width

    return wb


def main() -> int:
    git_counts = count_git_commits(REPO_ROOT)
    summary = build_summary(TASKS, git_counts)

    # The git figures above are a live measurement, not a timeless fact: they
    # will already be stale by the commit that adds this run's output (see
    # the coordinator note that caught this). Stamp the sha and date they
    # were measured at instead of pretending they hold forever.
    sha = get_git_sha(REPO_ROOT)
    measured_at = f"Git commit counts measured at {sha} on {date.today().isoformat()}."

    md = render_markdown(summary, TASKS, measured_at)
    MD_PATH.write_text(md, encoding="utf-8")
    print(f"Wrote {MD_PATH.relative_to(REPO_ROOT)} ({len(TASKS)} tasks, "
          f"{summary['total_hours']:g} hours, {summary['total_git']} commits, "
          f"{measured_at})")

    wb = render_xlsx(summary, TASKS, measured_at)
    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH.relative_to(REPO_ROOT)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
